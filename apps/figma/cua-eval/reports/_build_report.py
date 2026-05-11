"""Build a Google-Sheets-friendly XLSX and a one-pager PDF for the
2026-05-10 run of tasks 1-7 + 10 (Sonnet 4.5, --prompt-mode description,
no system prompt, keyboard allowed)."""
from __future__ import annotations

import glob
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, KeepTogether,
)


RUNS_ROOT = Path("/Users/v/Repos/figma-mock/apps/figma/cua-eval/runs")
OUT = Path("/Users/v/Repos/figma-mock/apps/figma/cua-eval/reports")


# (display_id, task_short_name, run_dir, status_label)
TASKS = [
    ("01", "Two-story house",          "20260510_201720", "done"),
    ("02", "Sunset stripe band",       "20260510_201730", "done"),
    ("03", "Radial flower with petals","20260510_203236", "done"),
    ("04", "Color hexagon ring",       "20260510_210418", "done"),
    ("05", "Plus-sign emblem",         "20260510_202047", "done"),
    ("06", "Asterisk burst",           "20260510_203303", "done"),
    ("07", "Layered mountain range",   "20260510_203315", "done"),
    ("08", "Layered water waves",      "20260510_210426", "done"),
    ("09", "12-color swatch grid",     None,              "skipped"),
    ("10", "Concentric squares",       "20260510_205934", "done"),
]


# Per-task curated failure-mode summary. These are derived from the rubric
# breakdowns and trajectories — kept terse for the one-pager.
FAILURE_MODES = {
    "01": "Polygon roof rotated 180° (apex pointing down). Polygon doesn't sit flush above body rectangle. One body rectangle drawn too narrow (width fraction 0.04). All other rubrics 100%.",
    "02": "Got 5 stripes correctly stacked at uniform size, but ALL fill colors wrong (color diff 0.35 on layer 1; sunset gradient palette not matched). Color rubric scored 0/0.40.",
    "03": "All 9 ellipses drawn with distinct colors, but the 8 outer petals not arranged radially: gap-angle deviation 9.8° (tol 15° passed) AND radius ratio 1.44 (need ~1.0) failed. Burned 104 turns trying to reposition.",
    "04": "6 squares created with 7 distinct colors, but NOT arranged in a ring: gap-angle deviation 22° (tol 10°) and radius spread 61px (tol 22°) both fail. Layout looks like a cluster, not a hexagon.",
    "06": "Created 8 lines, but endpoint cluster too weak — only 6/8 endpoints within 20px of shared center. Stroke color is some yellow but not the target gold (#d9a521).",
    "07": "Passed at 0.78. All shape, alignment, color, structure rubrics 100%. Only weak spot: `create_vector` event count = 0 (model used pen but the semantic log didn't record vector creations). Above threshold.",
    "08": "Passed at 0.78. 2 curved vectors with matching stroke weights, distinct stroke colors — all geometric/color rubrics 100%. Same `create_vector` event-count miss as task 07. Burned 145 turns to get there (target 40), pushing cost to $7.93 — most expensive single task.",
    "10": "All 4 concentric rectangles drawn correctly with alternating fills and proper nesting — base score 1.00. Failed only on efficiency: 45 turns vs 18 target → eff multiplier 0.63.",
}


def load_attempt(run_dir: str) -> tuple[dict, dict, dict] | None:
    if run_dir is None:
        return None
    attempts_p = RUNS_ROOT / run_dir / "attempts.json"
    if not attempts_p.is_file():
        return None
    with attempts_p.open() as f:
        attempts = json.load(f)
    if not attempts:
        return None
    a = attempts[0]
    score_p = Path(a["score_path"])
    with score_p.open() as f:
        score = json.load(f)
    # meta
    meta_glob = glob.glob(str(RUNS_ROOT / run_dir / "anthropic" / "task_*" / "attempt_1" / "meta.json"))
    with open(meta_glob[0]) as f:
        meta = json.load(f)
    return a, score, meta


def fmt_pct(x: float) -> str:
    return f"{x * 100:.0f}%"


def fmt_usd(x: float) -> str:
    return f"${x:.2f}"


# ─── Build XLSX ──────────────────────────────────────────────────────────────

def build_xlsx() -> Path:
    wb = openpyxl.Workbook()

    # Sheet 1: headline summary
    s1 = wb.active
    s1.title = "Summary"
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    pass_fill = PatternFill("solid", fgColor="D1FAE5")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    skip_fill = PatternFill("solid", fgColor="F3F4F6")
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Task #", "Name", "Status", "Pass", "Final score", "Base score",
        "Efficiency ×", "Turns", "Target turns", "Time (s)",
        "Input tok", "Output tok", "Cost (USD)", "Failure mode",
    ]
    s1.append(headers)
    for col, _ in enumerate(headers, start=1):
        c = s1.cell(row=1, column=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    total_cost = 0.0
    rows_data = []
    for tid, name, run_dir, status in TASKS:
        if run_dir is None:
            rows_data.append({
                "tid": tid, "name": name, "status": status, "skipped": True,
            })
            continue
        loaded = load_attempt(run_dir)
        if loaded is None:
            rows_data.append({
                "tid": tid, "name": name, "status": "missing", "skipped": True,
            })
            continue
        a, score, _meta = loaded
        cost = (a.get("cost_estimate") or {}).get("total_usd", 0.0) or 0.0
        total_cost += cost
        rows_data.append({
            "tid": tid, "name": name, "status": "done", "skipped": False,
            "passed": a["passed"], "final": a["final_score"], "base": a["base_score"],
            "eff": a["efficiency"], "turns": a["turns"],
            "target_turns": score["efficiency"]["target_turns"],
            "elapsed_s": a["elapsed_s"],
            "in_tok": (a["usage"] or {}).get("input_tokens", 0),
            "out_tok": (a["usage"] or {}).get("output_tokens", 0),
            "cost": cost,
            "failure": FAILURE_MODES.get(tid, "" if a["passed"] else "see score.json"),
        })

    for r_idx, r in enumerate(rows_data, start=2):
        if r["skipped"]:
            row = [r["tid"], r["name"], r["status"], "", "", "", "", "", "", "", "", "", "", ""]
            for c_idx, val in enumerate(row, start=1):
                cell = s1.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = skip_fill
                cell.border = border
            continue
        row = [
            r["tid"], r["name"], "done",
            "PASS" if r["passed"] else "FAIL",
            round(r["final"], 3), round(r["base"], 3),
            round(r["eff"], 3),
            r["turns"], r["target_turns"],
            round(r["elapsed_s"], 1),
            r["in_tok"], r["out_tok"],
            round(r["cost"], 4),
            r["failure"],
        ]
        for c_idx, val in enumerate(row, start=1):
            cell = s1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if c_idx == 4:
                cell.fill = pass_fill if r["passed"] else fail_fill
                cell.font = Font(bold=True, color=("065F46" if r["passed"] else "991B1B"))
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(c_idx == 14))

    # Totals row
    done_rows = [r for r in rows_data if not r["skipped"]]
    pass_count = sum(1 for r in done_rows if r["passed"])
    total_in = sum(r["in_tok"] for r in done_rows)
    total_out = sum(r["out_tok"] for r in done_rows)
    avg_score = (sum(r["final"] for r in done_rows) / len(done_rows)) if done_rows else 0
    tot_row = len(rows_data) + 2
    s1.cell(row=tot_row, column=2, value=f"Totals ({len(done_rows)} done, {pass_count} passed)").font = Font(bold=True)
    s1.cell(row=tot_row, column=5, value=round(avg_score, 3)).font = Font(bold=True)
    s1.cell(row=tot_row, column=11, value=total_in).font = Font(bold=True)
    s1.cell(row=tot_row, column=12, value=total_out).font = Font(bold=True)
    s1.cell(row=tot_row, column=13, value=round(total_cost, 4)).font = Font(bold=True)

    # Column widths
    widths = [8, 30, 10, 8, 12, 12, 12, 8, 12, 10, 12, 12, 12, 80]
    for i, w in enumerate(widths, start=1):
        s1.column_dimensions[get_column_letter(i)].width = w

    # Freeze headers
    s1.freeze_panes = "A2"

    # Sheet 2: rubric breakdown
    s2 = wb.create_sheet("Rubric breakdown")
    headers2 = ["Task #", "Name", "Rubric", "Score", "Max", "Percent", "Checks (passed/total)", "Notable fails"]
    s2.append(headers2)
    for col, _ in enumerate(headers2, start=1):
        c = s2.cell(row=1, column=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    row_i = 2
    for tid, name, run_dir, _status in TASKS:
        if run_dir is None:
            continue
        loaded = load_attempt(run_dir)
        if loaded is None:
            continue
        _a, score, _meta = loaded
        for rubric in score["rubrics"]:
            checks = rubric.get("checks") or []
            n_pass = sum(1 for c in checks if c["passed"])
            n_total = len(checks)
            fails = [c["message"].replace("\n", " ")[:140] for c in checks if not c["passed"]]
            fails_txt = " ; ".join(fails) if fails else ""
            pct = rubric["score"] / rubric["max_score"] if rubric["max_score"] else 0
            s2.append([
                tid, name, rubric["name"],
                round(rubric["score"], 3), round(rubric["max_score"], 2),
                fmt_pct(pct), f"{n_pass}/{n_total}", fails_txt,
            ])
            for col in range(1, 9):
                cell = s2.cell(row=row_i, column=col)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col == 8))
                if col == 6 and pct < 1.0:
                    cell.fill = fail_fill
            row_i += 1

    widths2 = [8, 24, 14, 10, 10, 10, 18, 100]
    for i, w in enumerate(widths2, start=1):
        s2.column_dimensions[get_column_letter(i)].width = w
    s2.freeze_panes = "A2"

    out_path = OUT / "tasks_1-10_summary.xlsx"
    wb.save(out_path)
    return out_path, rows_data, total_cost


# ─── Build PDF one-pager ─────────────────────────────────────────────────────

def build_pdf(rows_data, total_cost) -> Path:
    out_path = OUT / "tasks_1-10_summary.pdf"
    doc = SimpleDocTemplate(
        str(out_path), pagesize=letter,
        leftMargin=0.45 * inch, rightMargin=0.45 * inch,
        topMargin=0.5 * inch, bottomMargin=0.45 * inch,
    )
    styles = getSampleStyleSheet()
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=8.5, leading=11)
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=7.5, leading=9, textColor=colors.HexColor("#374151"))
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=15, leading=17, spaceAfter=2, textColor=colors.HexColor("#111827"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=10.5, leading=12, spaceBefore=6, spaceAfter=3, textColor=colors.HexColor("#374151"))
    subtle = ParagraphStyle("subtle", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.HexColor("#6B7280"))

    elements = []
    elements.append(Paragraph("Figma CUA Benchmark — Tasks 1–10", h1))
    elements.append(Paragraph(
        "Model: <b>claude-sonnet-4-5</b> &nbsp;|&nbsp; Run date: 2026-05-10 &nbsp;|&nbsp; "
        "k = 1 &nbsp;|&nbsp; threshold: final_score ≥ 0.7 &nbsp;|&nbsp; "
        "prompt-mode: description &nbsp;|&nbsp; no system prompt &nbsp;|&nbsp; keyboard allowed", subtle))
    elements.append(Spacer(1, 6))

    done_rows = [r for r in rows_data if not r["skipped"]]
    skipped_rows = [r for r in rows_data if r["skipped"]]
    pass_count = sum(1 for r in done_rows if r["passed"])
    avg_score = sum(r["final"] for r in done_rows) / len(done_rows) if done_rows else 0
    avg_eff = sum(r["eff"] for r in done_rows) / len(done_rows) if done_rows else 0
    avg_turns = sum(r["turns"] for r in done_rows) / len(done_rows) if done_rows else 0
    total_in = sum(r["in_tok"] for r in done_rows)
    total_out = sum(r["out_tok"] for r in done_rows)

    # Headline metric tiles
    tile_data = [
        ["pass@1", "mean score", "mean efficiency ×", "mean turns", "total cost"],
        [
            f"{pass_count}/{len(done_rows)}  ({pass_count/len(done_rows)*100:.0f}%)" if done_rows else "—",
            f"{avg_score:.3f}",
            f"{avg_eff:.3f}",
            f"{avg_turns:.0f}",
            f"${total_cost:.2f}",
        ],
    ]
    tile_t = Table(tile_data, colWidths=[1.5 * inch] * 5)
    tile_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F9FAFB")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("TOPPADDING", (0, 1), (-1, 1), 8),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
    ]))
    elements.append(tile_t)
    elements.append(Spacer(1, 8))

    # Per-task table
    elements.append(Paragraph("Per-task results", h2))
    tdata = [["#", "Task", "Pass", "Score", "Base", "Eff×", "Turns / target", "Time", "Cost", "Failure mode"]]
    for r in rows_data:
        if r["skipped"]:
            tdata.append([
                r["tid"], Paragraph(f"<i>{r['name']}</i>", small),
                r["status"][:8], "", "", "", "", "", "",
                Paragraph(f"<font color='#9CA3AF'><i>{r['status']}</i></font>", small),
            ])
        else:
            mark = "PASS" if r["passed"] else "FAIL"
            failure = r["failure"] or "—"
            tdata.append([
                r["tid"],
                Paragraph(r["name"], small),
                mark,
                f"{r['final']:.3f}",
                f"{r['base']:.2f}",
                f"{r['eff']:.2f}",
                f"{r['turns']} / {r['target_turns']}",
                f"{r['elapsed_s']:.0f}s",
                f"${r['cost']:.2f}",
                Paragraph(failure, small),
            ])

    col_widths = [0.28, 1.3, 0.45, 0.55, 0.5, 0.5, 0.85, 0.5, 0.5, 2.4]
    col_widths = [w * inch for w in col_widths]
    t = Table(tdata, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for ri, r in enumerate(rows_data, start=1):
        if r["skipped"]:
            style_cmds.append(("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#F9FAFB")))
        else:
            if r["passed"]:
                style_cmds.append(("BACKGROUND", (2, ri), (2, ri), colors.HexColor("#D1FAE5")))
                style_cmds.append(("TEXTCOLOR", (2, ri), (2, ri), colors.HexColor("#065F46")))
            else:
                style_cmds.append(("BACKGROUND", (2, ri), (2, ri), colors.HexColor("#FEE2E2")))
                style_cmds.append(("TEXTCOLOR", (2, ri), (2, ri), colors.HexColor("#991B1B")))
            style_cmds.append(("FONTNAME", (2, ri), (2, ri), "Helvetica-Bold"))
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)
    elements.append(Spacer(1, 6))

    # Patterns / takeaways
    elements.append(Paragraph("Patterns", h2))
    bullets = [
        "<b>Geometric arrangements are the dominant failure mode.</b> Radial placements (task 03 orb, task 04 wheel, task 06 star) all failed on positioning — the model gets shape counts and colors right, then can't lay them out in a ring or with matching radii.",
        "<b>Colors mismatch when palette is specific.</b> Task 02 (sunset gradient) had 5 stripes correctly sized and stacked but used wrong colors entirely — color rubric scored 0/0.40.",
        "<b>Efficiency is the silent killer.</b> Task 10 nailed all rubrics (base 1.0) but burned 45 turns vs 18 target → final 0.63, missed pass threshold. Task 01 likewise has 0.885 base but 0.70 efficiency multiplier.",
        "<b>Small-detail orientation errors.</b> Task 01 polygon roof rotated 180° (apex pointing down). Same class of error as not knowing canonical orientation.",
        "<b>When tasks are simple and few-shape, performance is excellent.</b> Task 05 (plus sign, 2 rectangles): 1.000 score in 7 turns. Task 07 (mountain range, pen tool): 0.778 in 41 turns.",
        "<b>Long-tail cost risk.</b> Task 08 passed but burned 145 turns and $7.93 — half the total spend across all 9 tasks. With no step cap, a single stubborn task can dominate budget. Consider re-introducing a cap (e.g. 80–100) for budget safety.",
    ]
    for b in bullets:
        elements.append(Paragraph("• " + b, body))
        elements.append(Spacer(1, 2))

    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f"<i>Generated from {len(done_rows)} attempts ({total_in:,} input tokens, "
        f"{total_out:,} output tokens, ${total_cost:.2f} spend). Source: "
        f"apps/figma/cua-eval/runs/202605*. Skipped: task 09 (not run yet).</i>",
        subtle))

    doc.build(elements)
    return out_path


if __name__ == "__main__":
    OUT.mkdir(parents=True, exist_ok=True)
    xlsx_path, rows, total_cost = build_xlsx()
    pdf_path = build_pdf(rows, total_cost)
    print(f"XLSX → {xlsx_path}")
    print(f"PDF  → {pdf_path}")
