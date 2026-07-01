#!/usr/bin/env python3
"""gen_ribbon_inventory.py — convert Microsoft's official Word ribbon control-identifier
workbook into the parity pipeline's ribbon STRUCTURE ground truth.

Source workbook: parity/oracle/wordcontrols-m365-current.xlsx
  (github.com/OfficeDev/office-fluent-ui-command-identifiers,
   "Microsoft 365 / Current Channel / wordcontrols.xlsx" — matches the locked parity
   target ADR-0006: M365 Word-for-Windows Current Channel.)

Output: parity/oracle/word_ribbon_inventory.json
  {
    "meta":     { source, channel, generated_by, row_count, counts_by_scope },
    "controls": [ { idMso, type, tabset, tab, group, parent, parent2, policyId }, ... ]
  }

Scope buckets (meta.counts_by_scope):
  core_tab        — the 10 main ribbon tabs (the clone's WC.RIBBON counterpart)
  contextual      — TabSet* contextual tabs (Table Tools, Picture Tools, ...)
  context_menu    — right-click menus
  backstage       — File/Backstage view
  qat             — Quick Access Toolbar defaults
  hidden          — "Not in the Ribbon" commands (keyboard/legacy/hidden features)
  other           — anything else

This file is the authoritative enumeration for the STRUCTURE parity axis: the clone's
ribbon-data.js is diffed against it (see the structure verifier), so missing buttons on
contextual tabs can no longer hide behind the clone's own subset being used as the spec.
Re-run after refreshing the workbook: python parity/tools/gen_ribbon_inventory.py
"""
import json
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "oracle" / "wordcontrols-m365-current.xlsx"
LABELS_TSV = ROOT / "oracle" / "_labels_mso.tsv"  # from enrich_labels_mso.ps1 (optional)
OUT = ROOT / "oracle" / "word_ribbon_inventory.json"


def scope_of(tabset):
    ts = tabset or ""
    if ts.startswith("TabSet"):
        return "contextual"
    if ts == "None (Core Tab)":
        return "core_tab"
    if ts == "None (Context Menu)":
        return "context_menu"
    if ts == "None (Backstage View)":
        return "backstage"
    if ts == "None (Quick Access Toolbar)":
        return "qat"
    if ts == "None (Not in the Ribbon)":
        return "hidden"
    return "other"


def main():
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl required: pip install openpyxl")
    if not XLSX.exists():
        sys.exit(f"workbook not found: {XLSX}")

    wb = load_workbook(XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Real-Word labels/enablement captured from the locked build via enrich_labels_mso.ps1.
    labels = {}
    if LABELS_TSV.exists():
        for line in LABELS_TSV.read_text(encoding="utf-8").splitlines():
            parts = line.split("\t")
            if len(parts) == 3 and parts[0]:
                labels[parts[0]] = {
                    "label": parts[1] or None,
                    "enabledBlankDoc": {"True": True, "False": False}.get(parts[2]),
                }

    controls = []
    for r in rows:
        idmso, ctype, tabset, tab, group, parent, parent2, policy = (r + (None,) * 9)[:8]
        if not idmso:
            continue
        mso = labels.get(str(idmso), {})
        controls.append({
            "idMso": str(idmso),
            "type": str(ctype) if ctype else None,
            "scope": scope_of(tabset),
            "tabset": str(tabset) if tabset else None,
            "tab": str(tab) if tab else None,
            "group": str(group) if group else None,
            "parent": str(parent) if parent else None,
            "parent2": str(parent2) if parent2 else None,
            "policyId": policy,
            "label": mso.get("label"),
            "enabledBlankDoc": mso.get("enabledBlankDoc"),
        })

    meta = {
        "source": "github.com/OfficeDev/office-fluent-ui-command-identifiers",
        "workbook": "Microsoft 365 / Current Channel / wordcontrols.xlsx",
        "local_copy": "parity/oracle/wordcontrols-m365-current.xlsx",
        "parity_target": "ADR-0006 M365 Word-for-Windows Current Channel x64 en-US",
        "generated_by": "parity/tools/gen_ribbon_inventory.py",
        "row_count": len(controls),
        "labels_from_real_word": bool(labels),
        "labeled_count": sum(1 for c in controls if c["label"]),
        "counts_by_scope": dict(Counter(c["scope"] for c in controls)),
        "counts_by_type": dict(Counter(c["type"] for c in controls)),
    }

    OUT.write_text(json.dumps({"meta": meta, "controls": controls}, indent=1), encoding="utf-8")
    print(f"wrote {OUT.relative_to(ROOT.parent)}: {len(controls)} controls")
    for k, v in sorted(meta["counts_by_scope"].items(), key=lambda kv: -kv[1]):
        print(f"  {v:5d}  {k}")


if __name__ == "__main__":
    main()
